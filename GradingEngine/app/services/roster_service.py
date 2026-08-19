"""
Exam roster (Excel) parsing, handwritten student-ID extraction, and
roster ↔ paper validation.
"""
from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.services.batch_upload import resolve_effective_batch_root
from app.services.ocr_service import process_student_answer

# Typical campus IDs e.g. IT22145976, SE12345678, or plain digit runs.
_STUDENT_ID_PATTERNS = [
    re.compile(
        r"(?i)(?:student\s*id|index\s*no\.?|index\s*number|reg(?:istration)?\s*no\.?|id)\s*[:\-#]?\s*"
        r"([A-Z]{1,4}\d{5,12})"
    ),
    re.compile(r"(?i)\b([A-Z]{1,3}\d{6,12})\b"),
    re.compile(r"(?i)\b(\d{7,12})\b"),
]

_ID_HEADER_ALIASES = {
    "student_id",
    "studentid",
    "student id",
    "id",
    "index",
    "index no",
    "index no.",
    "index number",
    "reg no",
    "registration no",
}
_NAME_HEADER_ALIASES = {"name", "student name", "student_name", "full name", "fullname"}
_DATE_HEADER_ALIASES = {"date", "exam date", "signed date", "attendance date"}


def normalize_student_id(value: Any) -> str:
    text = str(value or "").strip().upper()
    text = re.sub(r"[\s\-_/]+", "", text)
    return text


_FOLDER_ID_RE = re.compile(r"^(?:[A-Z]{1,4}\d{5,12}|\d{7,12})$")


def looks_like_student_id(value: Any) -> bool:
    """True when a folder/OCR token resembles a campus student ID."""
    return bool(_FOLDER_ID_RE.match(normalize_student_id(value)))


def extract_student_id_from_text(text: str) -> str | None:
    """Pull a student ID from OCR text (prefers labeled header lines)."""
    if not (text or "").strip():
        return None

    lines = [ln.strip() for ln in text.replace("\r", "\n").split("\n") if ln.strip()]
    header = "\n".join(lines[:25])
    full = "\n".join(lines)

    for blob in (header, full):
        for pattern in _STUDENT_ID_PATTERNS:
            match = pattern.search(blob)
            if match:
                return normalize_student_id(match.group(1))
    return None


def _norm_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def parse_roster_excel(file_bytes: bytes) -> dict:
    """
    Parse attendance Excel into roster students + duplicate_roster_ids.
    """
    from io import BytesIO

    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ValueError("Excel sheet is empty.")

    header_row = [_norm_header(c) for c in rows[0]]
    id_col = next((i for i, h in enumerate(header_row) if h in _ID_HEADER_ALIASES), None)
    name_col = next((i for i, h in enumerate(header_row) if h in _NAME_HEADER_ALIASES), None)
    date_col = next((i for i, h in enumerate(header_row) if h in _DATE_HEADER_ALIASES), None)

    if id_col is None:
        # Fallback: first column is student id
        id_col = 0

    students: list[dict] = []
    seen_order: list[str] = []
    counts: Counter[str] = Counter()

    for raw in rows[1:]:
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue
        cells = list(raw)
        sid = normalize_student_id(cells[id_col] if id_col < len(cells) else "")
        if not sid:
            continue
        name = ""
        if name_col is not None and name_col < len(cells) and cells[name_col] is not None:
            name = str(cells[name_col]).strip()
        date_val = None
        if date_col is not None and date_col < len(cells) and cells[date_col] is not None:
            date_val = str(cells[date_col]).strip()

        counts[sid] += 1
        if counts[sid] == 1:
            seen_order.append(sid)
            students.append(
                {
                    "student_id": sid,
                    "name": name,
                    **({"date": date_val} if date_val else {}),
                }
            )

    duplicate_roster_ids = sorted(sid for sid, n in counts.items() if n > 1)
    return {
        "students": students,
        "duplicate_roster_ids": duplicate_roster_ids,
        "uploaded_at": datetime.utcnow().isoformat() + "Z",
        "row_count": len(students),
    }


def list_paper_folders(batch_root: Path) -> list[str]:
    if not batch_root.is_dir():
        return []
    return sorted(
        p.name
        for p in batch_root.iterdir()
        if p.is_dir() and not p.name.startswith(".") and not p.name.startswith("_")
    )


def _first_page_file(paper_dir: Path) -> Path | None:
    files = []
    for entry in sorted(paper_dir.iterdir()):
        if not entry.is_file():
            continue
        lower = entry.name.lower()
        if lower.endswith((".png", ".jpg", ".jpeg", ".pdf", ".webp", ".tif", ".tiff", ".bmp")):
            files.append(entry)
    return files[0] if files else None


def id_scan_path(batch_dir: Path) -> Path:
    return batch_dir / "_id_scan.json"


def load_id_scan(batch_dir: Path) -> dict:
    path = id_scan_path(batch_dir)
    if not path.is_file():
        return {"papers": [], "scanned_at": None}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, dict):
            data.setdefault("papers", [])
            return data
    except (OSError, json.JSONDecodeError):
        pass
    return {"papers": [], "scanned_at": None}


def save_id_scan(batch_dir: Path, payload: dict) -> None:
    path = id_scan_path(batch_dir)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def effective_paper_id(paper: dict) -> str | None:
    """
    ID used for roster matching / grading:

    1. Manual override (lecturer fix)
    2. Folder name when it looks like a student ID
    3. OCR-extracted ID (fallback only)
    """
    manual = normalize_student_id(paper.get("manual_student_id"))
    if manual:
        return manual
    folder = normalize_student_id(paper.get("paper_key"))
    if looks_like_student_id(folder):
        return folder
    ocr = normalize_student_id(paper.get("ocr_student_id"))
    return ocr or None


async def scan_batch_student_ids(batch_dir: Path) -> dict:
    """
    Build per-paper ID records for roster validation.

    When the folder name already looks like a student ID, that is used and
    first-page OCR is skipped. Otherwise OCR the first page as a fallback.
    Preserves any existing manual_student_id overrides.
    """
    root = resolve_effective_batch_root(batch_dir)
    previous = {p.get("paper_key"): p for p in load_id_scan(root).get("papers") or [] if isinstance(p, dict)}

    papers: list[dict] = []
    for paper_key in list_paper_folders(root):
        paper_path = root / paper_key
        first = _first_page_file(paper_path)
        prev = previous.get(paper_key) or {}
        folder_id = normalize_student_id(paper_key)

        row: dict[str, Any] = {
            "paper_key": paper_key,
            "ocr_student_id": None,
            "manual_student_id": prev.get("manual_student_id") or None,
            "header_preview": "",
            "error": None,
            "id_source": None,
        }

        # Preferred path: folder name is already the student ID.
        if looks_like_student_id(folder_id):
            row["id_source"] = "folder"
            papers.append(row)
            continue

        if first is None:
            row["error"] = "No image/PDF found in folder."
            papers.append(row)
            continue

        try:
            # Student ID lives on the first page only — never OCR later pages for ID scan.
            text, _pages = await process_student_answer(str(first), max_pages=1)
            text = (text or "").strip()
            row["header_preview"] = "\n".join(text.splitlines()[:12])[:800]
            row["ocr_student_id"] = extract_student_id_from_text(text)
            row["id_source"] = "ocr" if row["ocr_student_id"] else None
            if not row["ocr_student_id"]:
                row["error"] = "Could not read student ID from paper header."
        except Exception as err:
            row["error"] = f"OCR failed: {err}"

        papers.append(row)

    payload = {
        "scanned_at": datetime.utcnow().isoformat() + "Z",
        "batch_root": str(root),
        "papers": papers,
    }
    save_id_scan(root, payload)
    return payload


def apply_manual_ids(batch_dir: Path, overrides: dict[str, str]) -> dict:
    """Set manual_student_id for paper_keys. Empty string clears override."""
    root = resolve_effective_batch_root(batch_dir)
    scan = load_id_scan(root)
    papers = list(scan.get("papers") or [])
    by_key = {p.get("paper_key"): p for p in papers if isinstance(p, dict)}

    for paper_key, raw_id in overrides.items():
        key = str(paper_key)
        if key not in by_key:
            by_key[key] = {
                "paper_key": key,
                "ocr_student_id": None,
                "manual_student_id": None,
                "header_preview": "",
                "error": None,
            }
            papers.append(by_key[key])
        normalized = normalize_student_id(raw_id)
        by_key[key]["manual_student_id"] = normalized or None
        if normalized:
            by_key[key]["error"] = None

    scan["papers"] = papers
    scan["updated_at"] = datetime.utcnow().isoformat() + "Z"
    save_id_scan(root, scan)
    return scan


def validate_roster_against_papers(roster: dict | None, scan: dict | None) -> dict:
    """
    Compare exam roster to scanned/manual paper IDs.
    """
    roster = roster or {}
    scan = scan or {}
    roster_students = [
        s for s in (roster.get("students") or []) if isinstance(s, dict) and s.get("student_id")
    ]
    roster_ids = [normalize_student_id(s["student_id"]) for s in roster_students]
    roster_name = {
        normalize_student_id(s["student_id"]): str(s.get("name") or "").strip()
        for s in roster_students
    }
    duplicate_roster_ids = [
        normalize_student_id(x) for x in (roster.get("duplicate_roster_ids") or [])
    ]

    papers = [p for p in (scan.get("papers") or []) if isinstance(p, dict)]
    paper_by_id: dict[str, list[dict]] = defaultdict(list)
    unreadable: list[dict] = []

    for paper in papers:
        eid = effective_paper_id(paper)
        if not eid:
            unreadable.append(paper)
            continue
        paper_by_id[eid].append(paper)

    rows: list[dict] = []
    matched_paper_keys: list[str] = []
    summary = Counter()

    # Roster-side rows
    for sid in roster_ids:
        name = roster_name.get(sid) or None
        hits = paper_by_id.get(sid) or []
        keys = [str(p.get("paper_key")) for p in hits]

        if sid in duplicate_roster_ids:
            status = "duplicate_roster"
        elif len(hits) == 0:
            status = "missing_paper"
        elif len(hits) > 1:
            status = "duplicate_paper"
        else:
            status = "matched"
            matched_paper_keys.append(keys[0])

        summary[status] += 1
        rows.append(
            {
                "student_id": sid,
                "name": name,
                "status": status,
                "paper_keys": keys,
            }
        )

    # Extra papers (OCR/manual ID not on roster)
    roster_set = set(roster_ids)
    for sid, hits in paper_by_id.items():
        if sid in roster_set:
            continue
        keys = [str(p.get("paper_key")) for p in hits]
        status = "duplicate_paper" if len(hits) > 1 else "extra_paper"
        # If duplicates off-roster, still report as duplicate_paper once + extras conceptually
        if len(hits) > 1:
            summary["duplicate_paper"] += 1
            rows.append(
                {
                    "student_id": sid,
                    "name": None,
                    "status": "duplicate_paper",
                    "paper_keys": keys,
                }
            )
        else:
            summary["extra_paper"] += 1
            rows.append(
                {
                    "student_id": sid,
                    "name": None,
                    "status": "extra_paper",
                    "paper_keys": keys,
                }
            )

    for paper in unreadable:
        summary["unreadable_id"] += 1
        rows.append(
            {
                "student_id": None,
                "name": None,
                "status": "unreadable_id",
                "paper_keys": [str(paper.get("paper_key"))],
                "error": paper.get("error"),
                "header_preview": paper.get("header_preview") or "",
            }
        )

    hard_blockers = (
        summary["duplicate_paper"]
        + summary["duplicate_roster"]
        + summary["unreadable_id"]
    )
    soft_warnings = summary["missing_paper"] + summary["extra_paper"]

    return {
        "summary": {
            "matched": summary["matched"],
            "missing_paper": summary["missing_paper"],
            "extra_paper": summary["extra_paper"],
            "duplicate_paper": summary["duplicate_paper"],
            "duplicate_roster": summary["duplicate_roster"],
            "unreadable_id": summary["unreadable_id"],
            "roster_count": len(roster_ids),
            "paper_count": len(papers),
        },
        "rows": rows,
        "matched_paper_keys": matched_paper_keys,
        "can_grade": hard_blockers == 0 and summary["matched"] > 0,
        "hard_blockers": hard_blockers,
        "soft_warnings": soft_warnings,
        "scanned_at": scan.get("scanned_at"),
        "roster_uploaded_at": roster.get("uploaded_at"),
    }
